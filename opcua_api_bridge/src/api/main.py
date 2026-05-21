"""
OPC UA REST API 桥接器 - 主应用入口
Author: WorkBuddy SRE
Version: 3.0.0

v3.0.0 改动：
- 适配自适应采集模式（client.py v6.0.0）
- 记录启动时间供健康检查使用
- API 层只从缓存读取
"""

import os
import logging
import time
import asyncio
import io
import math
from contextlib import asynccontextmanager
from datetime import datetime
from typing import Optional, Dict, Any, List

from fastapi import FastAPI, Depends, HTTPException, status, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
import structlog
from pydantic import BaseModel, Field
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from src.config.settings import settings
from src.monitoring.metrics import metrics_registry
from src.monitoring.health import HealthCheck
from src.opcua_client.client import OPCUAClient

# 配置结构化日志
structlog.configure(
    processors=[
        structlog.processors.add_log_level,
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.JSONRenderer()
    ]
)
logger = structlog.get_logger()

# OPC UA 客户端实例
opcua_client: Optional[OPCUAClient] = None

# 认证
security = HTTPBearer()

# 应用启动时间
_start_time = time.time()

# 应用生命周期管理
@asynccontextmanager
async def lifespan(app: FastAPI):
    """应用生命周期管理"""
    global opcua_client
    logger.info("opcua_bridge_starting", version="3.0.0")
    
    # 初始化 OPC UA 客户端并启动自适应采集
    opcua_client = OPCUAClient(
        endpoint=settings.OPCUA_ENDPOINT,
        username=settings.OPCUA_USERNAME,
        password=settings.OPCUA_PASSWORD,
        security_policy=None,
        security_mode=None
    )
    opcua_client._start_time = _start_time

    # 预注册默认采集节点（必须在 start() 之前，否则无推送触发退避）
    # 节点来源：与 reporter 报表一致，R301~R310 共 10 反应器，每家仅 F01
    _DEFAULT_NODE_IDS = [
        f"ns=1;s={n}" for n in [
            "FIT_05R301F01.PV","FIT_05R302F01.PV","FIT_05R303F01.PV",
            "FIT_05R304F01.PV","FIT_05R305F01.PV","FIT_05R306F01.PV",
            "FIT_05R307F01.PV","FIT_05R308F01.PV","FIT_05R309F01.PV",
            "FIT_05R310F01.PV",
            "FIQ_05R301F01.OUT","FIQ_05R302F01.OUT","FIQ_05R303F01.OUT",
            "FIQ_05R304F01.OUT","FIQ_05R305F01.OUT","FIQ_05R306F01.OUT",
            "FIQ_05R307F01.OUT","FIQ_05R308F01.OUT","FIQ_05R309F01.OUT",
            "FIQ_05R310F01.OUT",
        ]
    ]
    opcua_client.add_nodes(_DEFAULT_NODE_IDS)

    ok = await opcua_client.start()
    if ok:
        logger.info("opcua_adaptive_collect_started", endpoint=settings.OPCUA_ENDPOINT)
    else:
        logger.error("opcua_collect_start_failed")
    
    yield
    
    # 关闭时
    if opcua_client:
        try:
            await opcua_client.stop()
        except Exception:
            pass
    logger.info("opcua_bridge_shutdown")


# 创建 FastAPI 应用
app = FastAPI(
    title="OPC UA REST API Bridge",
    description="轻量级 OPC UA 到 REST API 的桥接器（自适应采集模式）",
    version="3.0.0",
    lifespan=lifespan,
    docs_url="/docs" if settings.DEBUG else None,
    redoc_url="/redoc" if settings.DEBUG else None
)

# CORS 中间件
app.add_middleware(
    CORSMiddleware,
    allow_origins=settings.ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 全局异常处理
@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """全局异常处理"""
    logger.error("unexpected_error", path=request.url.path, error=str(exc), exc_info=True)
    return JSONResponse(
        status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
        content={"detail": "Internal server error"}
    )

# 健康检查路由
@app.get("/health", tags=["监控"])
async def health_check():
    """系统健康检查"""
    if not opcua_client:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content={"status": "unhealthy", "detail": "客户端未初始化"}
        )

    health = HealthCheck(opcua_client)
    status_result = await health.check_all()
    
    if status_result["status"] == "healthy":
        return status_result
    else:
        return JSONResponse(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            content=status_result
        )

# 节点管理路由
@app.get("/api/v1/nodes", tags=["节点管理"])
async def get_nodes(
    node_id: Optional[str] = Query(None, description="起始节点ID，不填则从 Objects 根节点开始"),
    namespace: Optional[int] = Query(None, description="命名空间过滤，如 1 只返回 ns=1 的节点"),
    recursive: bool = Query(False, description="是否递归（默认关闭，数据量大时容易超时）"),
    limit: int = Query(200, ge=1, le=1000, description="单次返回最大节点数"),
    offset: int = Query(0, ge=0, description="分页偏移量")
):
    """
    获取 OPC UA 节点列表（分页）
    - 临时短连接浏览，用完即释放 Session
    - 不传 node_id 默认从 Objects 节点（ns=0;i=85）开始
    """
    if not opcua_client:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="OPC UA 客户端未初始化"
        )

    try:
        nodes = await opcua_client.browse_nodes(
            node_id=node_id,
            namespace=namespace,
            recursive=recursive,
            max_nodes=limit,
            offset=offset
        )
        return {"nodes": nodes, "count": len(nodes), "offset": offset, "limit": limit}
    except Exception as e:
        logger.error("get_nodes_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"获取节点失败: {str(e)}"
        )

# 读取节点值
@app.get("/api/v1/nodes/{node_id}/value", tags=["数据读取"])
async def read_node_value(node_id: str):
    """
    读取指定节点的当前值（从缓存读取，并注册到采集列表）
    """
    if not opcua_client:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="OPC UA 客户端未连接"
        )
    
    try:
        value = await opcua_client.read_value(node_id)
        return {
            "node_id": node_id,
            "value": value["value"],
            "timestamp": value["timestamp"],
            "quality": value["quality"],
            "data_type": value["data_type"]
        }
    except Exception as e:
        logger.error("read_node_value_failed", node_id=node_id, error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"读取节点值失败: {str(e)}"
        )

# 批量读取节点值
@app.post("/api/v1/nodes/batch-read", tags=["数据读取"])
async def batch_read_values(node_ids: list[str]):
    """
    批量读取多个节点的值（从缓存读取，并注册到采集列表）。
    首次请求时节点可能尚未采集到数据，订阅推送后即可获取。
    """
    if not opcua_client:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="OPC UA 客户端未连接"
        )

    try:
        results = await opcua_client.batch_read(node_ids)
        return {"results": results, "count": len(results)}
    except Exception as e:
        logger.error("batch_read_failed", error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"批量读取失败: {str(e)}"
        )

# 查看订阅缓存状态
@app.get("/api/v1/cache/stats", tags=["监控"])
async def cache_stats():
    """查看订阅缓存统计"""
    if not opcua_client:
        raise HTTPException(status_code=503, detail="OPC UA 客户端未连接")
    return {
        "cache_size": opcua_client.get_cache_size(),
        "registered_nodes": opcua_client.get_subscribed_count(),
        "cache_freshness_seconds": opcua_client.get_cache_freshness(),
        "connected": opcua_client.is_connected(),
        "collect_status": opcua_client.get_collect_status(),
        "yielded": opcua_client.is_yielded(),
    }

# 读取历史数据
@app.get("/api/v1/nodes/{node_id}/history", tags=["历史数据"])
async def read_history(
    node_id: str,
    start_time: str = Query(..., description="开始时间 (ISO格式)"),
    end_time: str = Query(..., description="结束时间 (ISO格式)"),
    max_points: int = Query(1000, description="最大数据点数")
):
    """读取节点的历史数据"""
    if not opcua_client:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="OPC UA 客户端未连接"
        )
    
    try:
        history = await opcua_client.read_history(
            node_id=node_id,
            start_time=start_time,
            end_time=end_time,
            max_points=max_points
        )
        return {
            "node_id": node_id,
            "history": history,
            "count": len(history)
        }
    except Exception as e:
        logger.error("read_history_failed", node_id=node_id, error=str(e))
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"读取历史数据失败: {str(e)}"
        )

# Dashboard 看板
@app.get("/dashboard", tags=["可视化"])
async def dashboard():
    """OPC UA 可视化看板"""
    from fastapi.responses import FileResponse
    import os
    dashboard_path = os.path.join(os.path.dirname(__file__), "..", "..", "dashboard.html")
    return FileResponse(dashboard_path)

# Prometheus 指标端点
@app.get("/metrics", tags=["监控"])
async def metrics():
    """Prometheus 指标端点"""
    from prometheus_client import generate_latest, CONTENT_TYPE_LATEST
    from starlette.responses import Response
    
    return Response(
        content=generate_latest(metrics_registry),
        media_type=CONTENT_TYPE_LATEST
    )

# 根路由
@app.get("/")
async def root():
    """API 根路径"""
    return {
        "name": "OPC UA REST API Bridge",
        "version": "3.0.0",
        "mode": "adaptive_collect",
        "collect_status": opcua_client.get_collect_status() if opcua_client else "not_initialized",
        "status": "operational" if opcua_client else "degraded",
        "docs": "/docs" if settings.DEBUG else None,
        "health": "/health"
    }

# ─── 批量历史查询 + Excel 导出 ──────────────────────────

class BatchHistoryRequest(BaseModel):
    node_ids: List[str] = Field(..., description="节点ID列表，如 ['ns=1;s=FIT_05R301F01.PV']")
    start_time: str = Field(..., description="起始时间 ISO格式")
    end_time: str = Field(..., description="结束时间 ISO格式")
    interval_seconds: int = Field(60, ge=1, le=3600, description="采样间隔（秒），默认60")

class BatchHistoryResponse(BaseModel):
    series: List[Dict[str, Any]]
    interval_seconds: int
    total_points: int

def _sample_data(rows: List[Dict], interval_seconds: int, start_time: str, end_time: str) -> List[Dict]:
    """对原始数据按时间间隔采样，取每个窗口的第一个值"""
    if not rows or interval_seconds <= 0:
        return rows

    start_ts = datetime.fromisoformat(start_time).timestamp()
    end_ts = datetime.fromisoformat(end_time).timestamp()
    sampled = []

    for row in rows:
        try:
            ts = datetime.fromisoformat(row["timestamp"]).timestamp()
        except (ValueError, TypeError):
            continue
        if ts < start_ts or ts > end_ts:
            continue
        bucket = int((ts - start_ts) / interval_seconds)
        if not sampled or bucket != sampled[-1].get("_bucket", -1):
            row["_bucket"] = bucket
            sampled.append(row)

    # 清理内部字段
    for r in sampled:
        r.pop("_bucket", None)
    return sampled


@app.post("/api/v1/history/query", response_model=BatchHistoryResponse, tags=["历史数据"])
async def batch_history_query(req: BatchHistoryRequest):
    """
    批量查询多节点历史数据（从 SQLite），支持时间采样。
    - 返回每个节点按 interval_seconds 采样后的序列
    """
    if not opcua_client or not opcua_client.history_db:
        raise HTTPException(status_code=503, detail="历史数据库未就绪")

    series = []
    total = 0

    for node_id in req.node_ids:
        rows = await opcua_client.history_db.read_history(
            node_id=node_id,
            start_time=req.start_time,
            end_time=req.end_time,
            max_points=50000,
        )
        sampled = _sample_data(rows, req.interval_seconds, req.start_time, req.end_time)
        series.append({
            "node_id": node_id,
            "display_name": node_id.split("s=")[-1] if "s=" in node_id else node_id,
            "points": [{"timestamp": r["timestamp"], "value": r["value"]} for r in sampled],
            "count": len(sampled),
        })
        total += len(sampled)

    return {"series": series, "interval_seconds": req.interval_seconds, "total_points": total}


@app.post("/api/v1/history/export", tags=["历史数据"])
async def batch_history_export(req: BatchHistoryRequest):
    """
    批量查询历史数据并导出为 Excel 文件。
    - 参数与 /api/v1/history/query 一致
    - 返回 .xlsx 文件流
    """
    if not opcua_client or not opcua_client.history_db:
        raise HTTPException(status_code=503, detail="历史数据库未就绪")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "历史数据报表"

    # 样式
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    # 表头
    ws.cell(row=1, column=1, value="时间").font = header_font
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).alignment = header_align
    ws.cell(row=1, column=1).border = thin_border

    col_idx = 2
    for node_id in req.node_ids:
        display = node_id.split("s=")[-1] if "s=" in node_id else node_id
        cell = ws.cell(row=1, column=col_idx, value=display)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        col_idx += 1

    # 数据：按固定时间桶对齐（解决不同节点时间戳毫秒差异导致匹配失败的问题）
    start_ts = datetime.fromisoformat(req.start_time).timestamp()
    end_ts = datetime.fromisoformat(req.end_time).timestamp()
    interval = req.interval_seconds
    # 生成固定时间桶序列
    bucket_ts_list = []
    t = start_ts
    while t <= end_ts:
        bucket_ts_list.append(t)
        t += interval

    # 收集各节点数据并分配桶
    all_series = {}
    for node_id in req.node_ids:
        rows = await opcua_client.history_db.read_history(
            node_id=node_id, start_time=req.start_time,
            end_time=req.end_time, max_points=50000,
        )
        # 按桶分配：每个桶取最新一条数据
        bucket_vals = {}
        for row in rows:
            try:
                ts = datetime.fromisoformat(row["timestamp"]).timestamp()
            except (ValueError, TypeError):
                continue
            if ts < start_ts or ts > end_ts:
                continue
            bucket_idx = int((ts - start_ts) / interval)
            # 取桶内最接近桶边界的数据
            bucket_vals[bucket_idx] = row["value"]
        all_series[node_id] = bucket_vals

    for row_idx, bucket_ts in enumerate(bucket_ts_list, start=2):
        bucket_idx = row_idx - 2
        # 格式化显示时间（去掉毫秒）
        label = datetime.fromtimestamp(bucket_ts).strftime("%Y-%m-%d %H:%M:%S")
        ws.cell(row=row_idx, column=1, value=label).border = thin_border
        for col_offset, node_id in enumerate(req.node_ids):
            val = all_series.get(node_id, {}).get(bucket_idx)
            cell = ws.cell(row=row_idx, column=2 + col_offset, value=val)
            cell.border = thin_border
            if val is not None:
                cell.number_format = '0.00'

    # 调整列宽
    ws.column_dimensions["A"].width = 22
    for i in range(len(req.node_ids)):
        col_letter = openpyxl.utils.get_column_letter(2 + i)
        ws.column_dimensions[col_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"opcua_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# 启动说明
if __name__ == "__main__":
    import uvicorn
    
    logger.info("starting_opcua_bridge")
    uvicorn.run(
        "src.api.main:app",
        host=settings.HOST,
        port=settings.PORT,
        reload=settings.DEBUG,
        log_level="info"
    )

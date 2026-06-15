#!/usr/bin/env python3
"""
磺化釜搅拌电机电流周报 - Excel 版
- 分析上一完整周（周一 00:00 ~ 周日 23:59）的数据
- 输出 Excel：Sheet1 汇总统计 + 数据透视，之后每台电机一个 Sheet（运行数据 + 分时统计）
- 不生成趋势图，纯数据表格
- 停机阈值：基于直方图谷底检测，每台电机独立计算（同 motor_predictive_maintenance_report.py）
用法:
  python motor_current_excel_report.py               # 分析上一完整周
  python motor_current_excel_report.py --this-week   # 分析本周（周一~今天）
  python motor_current_excel_report.py --start 2026-06-01 --end 2026-06-07
"""

import argparse
import sqlite3
import os
import sys
from datetime import datetime, timedelta

# ── 必须先把项目根加到 sys.path，才能 import settings ─────────────────────────
sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
except ImportError:
    print("❌ 缺少 openpyxl，请运行：pip install openpyxl")
    sys.exit(1)


# ── 稳态数据提取 + 运行时长（与 motor_predictive_maintenance_report.py 一致）───

STOP_CURRENT_A = 1.0
START_CONSECUTIVE = 10
SURGE_REMOVE = 50
STOP_CONSECUTIVE = 10
PRE_STOP_REMOVE = 70
RUNTIME_START_N = 5
RUNTIME_STOP_N = 5


def extract_steady_segments(data_all):
    """三阶段状态机提取稳态数据。规则与预测性维护报告完全一致。

    停机→启动: 连续10点>1A, 启动点=第10个
    浪涌剔除: 启动点后50点
    稳态: 启动点+50 到 停机点-70
    稳态→停机: 连续10点<1A, 停机点=第10个
    """
    STOPPED, STARTING, STEADY = 0, 1, 2
    state = STOPPED
    above_count = 0; below_count = 0; cnt_since_start = 0
    steady_start_idx = None
    steady_ranges = []

    for i, d in enumerate(data_all):
        v = d["value"]
        if v is None: continue
        if state == STOPPED:
            if v > STOP_CURRENT_A:
                above_count += 1
                if above_count >= START_CONSECUTIVE:
                    state = STARTING; cnt_since_start = 0; above_count = 0
            else: above_count = 0
        elif state == STARTING:
            cnt_since_start += 1
            if cnt_since_start >= SURGE_REMOVE:
                state = STEADY; steady_start_idx = i + 1; below_count = 0
        elif state == STEADY:
            if v < STOP_CURRENT_A:
                below_count += 1
                if below_count >= STOP_CONSECUTIVE:
                    steady_end_idx = i - PRE_STOP_REMOVE
                    if steady_start_idx is not None and steady_end_idx >= steady_start_idx:
                        steady_ranges.append((steady_start_idx, steady_end_idx))
                    state = STOPPED; below_count = 0; steady_start_idx = None
            else: below_count = 0

    if state == STEADY and steady_start_idx is not None:
        steady_ranges.append((steady_start_idx, len(data_all) - 1))

    result = []
    for start, end in steady_ranges:
        for j in range(start, end + 1):
            if j < len(data_all) and data_all[j]["value"] is not None:
                result.append(data_all[j])
    return result


def calc_runtime(data_all):
    """计算累计运行时长。连续5点>1A开始计时, 连续5点<1A停止计时。"""
    RUNNING, STOPPED = "running", "stopped"
    state = STOPPED
    above_count = 0; below_count = 0
    run_start_ts = None
    total_seconds = 0.0

    for d in data_all:
        v = d["value"]
        if v is None: continue
        try: ts = datetime.fromisoformat(d["timestamp"])
        except (ValueError, TypeError): continue

        if state == STOPPED:
            if v > STOP_CURRENT_A:
                above_count += 1
                if above_count >= RUNTIME_START_N:
                    state = RUNNING; run_start_ts = ts; above_count = 0
            else: above_count = 0
        elif state == RUNNING:
            if v < STOP_CURRENT_A:
                below_count += 1
                if below_count >= RUNTIME_STOP_N:
                    state = STOPPED
                    if run_start_ts is not None:
                        total_seconds += max(0.0, (ts - run_start_ts).total_seconds())
                    run_start_ts = None; below_count = 0
            else: below_count = 0

    if state == RUNNING and run_start_ts is not None:
        try:
            last_ts = datetime.fromisoformat(data_all[-1]["timestamp"])
            total_seconds += max(0.0, (last_ts - run_start_ts).total_seconds())
        except (ValueError, TypeError, IndexError): pass

    return round(total_seconds / 3600.0, 1)


# ── 配置 ──────────────────────────────────────────────────────────────────────
DB_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "data", "history.db"
)

MOTOR_NODES = [
    "ns=1;s=IIAS_05A102.PV",
    "ns=1;s=IIAS_05A103.PV",
    "ns=1;s=IIAS_05A104.PV",
    "ns=1;s=IIAS_05A105.PV",
    "ns=1;s=IIAS_05A106.PV",
    "ns=1;s=IIAS_05A107.PV",
    "ns=1;s=IIAS_05A108.PV",
    "ns=1;s=IIAS_05A109.PV",
    "ns=1;s=IIAS_05A110.PV",
    "ns=1;s=IIAS_05A111.PV",
]


REPORT_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "..", "reports"
)

# ── 样式常量 ──────────────────────────────────────────────────────────────────
CLR_HEADER    = "1F4E79"   # 深蓝（汇总表头）
CLR_HEADER2   = "2E75B6"   # 中蓝（电机表头）
CLR_SUB       = "BDD7EE"   # 浅蓝（分区小标题）
CLR_PIVOT_HDR = "375623"   # 深绿（透视表头）
CLR_PIVOT_SUB = "E2EFDA"   # 浅绿（透视偶数行）
CLR_ALT       = "EBF3FB"   # 交替行背景
CLR_WARN      = "FFE699"   # 黄色警告
CLR_ALERT     = "FF7043"   # 橙红色告警
CLR_WHITE     = "FFFFFF"
FONT_NAME     = "微软雅黑"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(name=FONT_NAME, bold=bold, color=color, size=size)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _left():
    return Alignment(horizontal="left", vertical="center")

def _safe_table(node_id: str) -> str:
    return "h_" + node_id.replace("=", "_eq_").replace(";", "_sc_").replace(".", "_dot_")

def display_name(node_id: str) -> str:
    if ";s=" in node_id:
        return node_id.split(";s=", 1)[-1]
    return node_id

# ── 数据读取 ──────────────────────────────────────────────────────────────────
def read_data(db, node_id, start_time, end_time_excl):
    table = _safe_table(node_id)
    cur = db.cursor()
    cur.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,))
    if not cur.fetchone():
        return []
    cur.execute(
        f"SELECT timestamp, value FROM [{table}] "
        f"WHERE timestamp >= ? AND timestamp < ? ORDER BY timestamp ASC",
        (start_time, end_time_excl)
    )
    return [{"timestamp": r[0], "value": r[1]} for r in cur.fetchall() if r[1] is not None]


def compute_stats(data_all, data_running):
    """计算运行统计。

    参数:
      data_all:     全量数据（含停机），用于运行时长计算
      data_running: 已提取的稳态运行数据（由 extract_steady_segments 产出）
    """
    total = len(data_all)
    running = len(data_running)
    stop_pct = round((total - running) / total * 100, 1) if total > 0 else None
    run_pct  = round(running / total * 100, 1)           if total > 0 else None

    if not data_running:
        return {
            "count": 0, "total": total,
            "avg": None, "max": None, "min": None, "std": None,
            "runtime_h": 0, "stop_pct": stop_pct, "run_pct": run_pct,
            "data_start": None, "data_end": None,
        }

    vals = [d["value"] for d in data_running]
    n    = len(vals)
    avg  = sum(vals) / n
    mx   = max(vals)
    mn   = min(vals)
    std  = (sum((v - avg) ** 2 for v in vals) / n) ** 0.5

    # 累计运行时长：规则二（宽松，连续5点>1A计时，连续5点<1A停时）
    cum_h = calc_runtime(data_all)

    return {
        "count": n, "total": total,
        "avg": round(avg, 2), "max": round(mx, 2), "min": round(mn, 2), "std": round(std, 2),
        "runtime_h": cum_h,
        "stop_pct": stop_pct, "run_pct": run_pct,
        "data_start": data_running[0]["timestamp"][:19],
        "data_end":   data_running[-1]["timestamp"][:19],
    }


def hourly_stats(data_running):
    """按小时统计：均值/最大/最小/数据点数"""
    buckets = {}
    for d in data_running:
        try:
            ts = datetime.fromisoformat(d["timestamp"])
        except Exception:
            continue
        h = ts.strftime("%Y-%m-%d %H:00")
        if h not in buckets:
            buckets[h] = []
        buckets[h].append(d["value"])
    rows = []
    for h in sorted(buckets):
        vs = buckets[h]
        n  = len(vs)
        avg = round(sum(vs) / n, 2)
        rows.append({
            "小时": h,
            "均值 (A)": avg,
            "最大值 (A)": round(max(vs), 2),
            "最小值 (A)": round(min(vs), 2),
            "数据点数": n,
        })
    return rows


def daily_stats(data_running):
    """按天统计：均值/最大/最小/运行时长（用小时桶累加）"""
    day_buckets = {}
    for d in data_running:
        try:
            ts = datetime.fromisoformat(d["timestamp"])
        except Exception:
            continue
        day = ts.strftime("%Y-%m-%d")
        if day not in day_buckets:
            day_buckets[day] = []
        day_buckets[day].append(d["value"])
    rows = []
    for day in sorted(day_buckets):
        vs = day_buckets[day]
        n  = len(vs)
        avg = round(sum(vs) / n, 2)
        rows.append({
            "日期": day,
            "均值 (A)": avg,
            "最大值 (A)": round(max(vs), 2),
            "最小值 (A)": round(min(vs), 2),
            "数据点数": n,
        })
    return rows


# ── Excel 写入工具函数 ─────────────────────────────────────────────────────────
def _set_col_width(ws, col_idx, width):
    ws.column_dimensions[get_column_letter(col_idx)].width = width

def _write_header_row(ws, row_num, headers, bg_hex, font_color="FFFFFF", start_col=1):
    for ci, h in enumerate(headers, start=start_col):
        c = ws.cell(row=row_num, column=ci, value=h)
        c.fill      = _fill(bg_hex)
        c.font      = _font(bold=True, color=font_color)
        c.alignment = _center()
        c.border    = _border()

def _write_data_row(ws, row_num, values, alt=False, start_col=1, bold_first=False):
    bg = CLR_ALT if alt else CLR_WHITE
    for ci, v in enumerate(values, start=start_col):
        c = ws.cell(row=row_num, column=ci, value=v)
        c.fill      = _fill(bg)
        c.font      = _font(bold=(bold_first and ci == start_col))
        c.alignment = _center() if ci > start_col else _left()
        c.border    = _border()

def _section_title(ws, row_num, title, col_span, bg_hex=CLR_SUB, start_col=1):
    c = ws.cell(row=row_num, column=start_col, value=title)
    c.fill      = _fill(bg_hex)
    c.font      = _font(bold=True, size=10)
    c.alignment = _left()
    c.border    = _border()
    if col_span > 1:
        ws.merge_cells(
            start_row=row_num, start_column=start_col,
            end_row=row_num, end_column=start_col + col_span - 1
        )


# ── Sheet 1：汇总统计 + 横向透视 ─────────────────────────────────────────────
def build_summary_sheet(wb, results, period_desc, generated_at):
    ws = wb.create_sheet("汇总统计")
    ws.freeze_panes = "B4"

    # ── 标题行 ──
    ws.merge_cells("A1:J1")
    t = ws["A1"]
    t.value     = f"磺化釜搅拌电机电流周报 — {period_desc}"
    t.font      = Font(name=FONT_NAME, bold=True, size=14, color=CLR_WHITE)
    t.fill      = _fill(CLR_HEADER)
    t.alignment = _center()
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    sub = ws["A2"]
    sub.value     = (
        f"生成时间：{generated_at}    数据来源：OPC UA 历史数据库    "
        f"停机判定：固定 1.0A | 稳态提取：三阶段状态机"
    )
    sub.font      = _font(color="595959", size=9)
    sub.fill      = _fill("DDEEFF")
    sub.alignment = _center()
    ws.row_dimensions[2].height = 18

    # ── 表头 ──
    headers = ["电机位号", "均值 (A)", "最大值 (A)", "最小值 (A)",
               "标准差 (A)", "累计运行 (h)", "稳态数据点", "总数据点数",
               "停机占比 (%)", "运行占比 (%)"]
    _write_header_row(ws, 3, headers, CLR_HEADER)
    ws.row_dimensions[3].height = 20

    # ── 数据行 ──
    all_avgs = []
    for ri, r in enumerate(results):
        s    = r["stats"]
        name = display_name(r["node_id"])
        alt  = (ri % 2 == 1)
        avg_v = s["avg"] if s["avg"] is not None else ""
        # 异常着色：停机占比 >= 50% 标黄；avg 超过 40A 标橙红
        row_vals = [
            name,
            avg_v,
            s["max"]       if s["max"] is not None else "",
            s["min"]       if s["min"] is not None else "",
            s["std"]       if s["std"] is not None else "",
            s["runtime_h"],
            s["count"],
            s["total"],
            s["stop_pct"]  if s["stop_pct"] is not None else "",
            s["run_pct"]   if s["run_pct"]  is not None else "",
        ]
        row_idx = ri + 4
        bg = CLR_ALT if alt else CLR_WHITE
        if s["stop_pct"] is not None and s["stop_pct"] >= 50:
            bg = CLR_WARN
        if s["avg"] is not None and s["avg"] > 40:
            bg = CLR_ALERT
        for ci, v in enumerate(row_vals, start=1):
            c = ws.cell(row=row_idx, column=ci, value=v)
            c.fill      = _fill(bg)
            c.font      = _font(bold=(ci == 1))
            c.alignment = _left() if ci == 1 else _center()
            c.border    = _border()
        if s["avg"] is not None:
            all_avgs.append(s["avg"])

    last_data_row = len(results) + 3
    # 汇总行
    summary_row = last_data_row + 1
    _write_header_row(ws, summary_row, ["机队均值", round(sum(all_avgs)/len(all_avgs), 2) if all_avgs else "",
                                         "", "", "", "", "", "", "", ""],
                       CLR_HEADER2)

    # ── 透视表：按天 × 电机 的均值矩阵 ──
    pivot_start = summary_row + 3
    _section_title(ws, pivot_start, "📊  日均值透视（按天 × 电机，单位 A）",
                   len(results) + 1, bg_hex=CLR_PIVOT_HDR)
    ws.cell(row=pivot_start, column=1).font = _font(bold=True, color=CLR_WHITE)

    # 收集所有日期
    all_days = set()
    day_motor = {}  # (day, motor) -> avg
    for r in results:
        name = display_name(r["node_id"])
        ds = daily_stats(r["data_running"])
        for row in ds:
            d = row["日期"]
            all_days.add(d)
            day_motor[(d, name)] = row["均值 (A)"]
    sorted_days = sorted(all_days)
    motor_names = [display_name(r["node_id"]) for r in results]

    # 透视表头
    piv_hdr = ["日期"] + motor_names
    _write_header_row(ws, pivot_start + 1, piv_hdr, CLR_PIVOT_HDR, start_col=1)

    for di, day in enumerate(sorted_days):
        prow = pivot_start + 2 + di
        bg = CLR_PIVOT_SUB if di % 2 == 0 else CLR_WHITE
        c = ws.cell(row=prow, column=1, value=day)
        c.fill = _fill(bg); c.font = _font(bold=True); c.alignment = _center(); c.border = _border()
        for mi, mn in enumerate(motor_names, start=2):
            val = day_motor.get((day, mn), "")
            c2 = ws.cell(row=prow, column=mi, value=val)
            c2.fill = _fill(bg); c2.font = _font(); c2.alignment = _center(); c2.border = _border()

    # ── 列宽 ──
    col_widths = [20, 12, 12, 12, 12, 14, 18, 14, 14, 12]
    for ci, w in enumerate(col_widths, start=1):
        _set_col_width(ws, ci, w)
    # 透视表列宽（电机名短，10 够了）
    for ci in range(2, len(motor_names) + 2):
        _set_col_width(ws, ci, 14)


# ── 每台电机 Sheet ─────────────────────────────────────────────────────────────
def build_motor_sheet(wb, r):
    name     = display_name(r["node_id"])
    ws       = wb.create_sheet(name[:31])  # Sheet 名最多 31 字符
    ws.freeze_panes = "A4"

    data_running = r["data_running"]
    data_all     = r["data_all"]
    s            = r["stats"]

    # ── 标题 ──
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value     = f"电机 {name} — 电流分析（稳态数据，停机判定 1.0A）"
    t.font      = Font(name=FONT_NAME, bold=True, size=12, color=CLR_WHITE)
    t.fill      = _fill(CLR_HEADER2)
    t.alignment = _center()
    ws.row_dimensions[1].height = 24

    # ── 统计摘要（关键指标 2 行） ──
    kpi_labels = ["均值 (A)", "最大值 (A)", "最小值 (A)", "标准差 (A)",
                  "累计运行 (h)", "停机占比 (%)", "运行数据点", "总数据点"]
    kpi_vals = [
        s["avg"]       if s["avg"]       is not None else "无数据",
        s["max"]       if s["max"]       is not None else "—",
        s["min"]       if s["min"]       is not None else "—",
        s["std"]       if s["std"]       is not None else "—",
        s["runtime_h"],
        s["stop_pct"]  if s["stop_pct"]  is not None else "—",
        s["count"],
        s["total"],
    ]
    _write_header_row(ws, 2, kpi_labels, CLR_SUB, font_color="1F4E79")
    _write_data_row(ws, 3, kpi_vals)
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 18

    cur_row = 5

    # ── 日均值分析 ──
    _section_title(ws, cur_row, "📅  日均值分析", 5)
    cur_row += 1
    day_headers = ["日期", "均值 (A)", "最大值 (A)", "最小值 (A)", "数据点数"]
    _write_header_row(ws, cur_row, day_headers, CLR_HEADER2)
    cur_row += 1
    day_rows = daily_stats(data_running)
    if not day_rows:
        ws.cell(row=cur_row, column=1, value="本周无运行数据").alignment = _left()
        cur_row += 1
    else:
        for di, dr in enumerate(day_rows):
            vals = [dr["日期"], dr["均值 (A)"], dr["最大值 (A)"], dr["最小值 (A)"], dr["数据点数"]]
            _write_data_row(ws, cur_row, vals, alt=(di % 2 == 1))
            cur_row += 1

    cur_row += 1  # 空行

    # ── 分时（小时）统计 ──
    _section_title(ws, cur_row, "🕐  分时统计（按小时）", 5)
    cur_row += 1
    _write_header_row(ws, cur_row, ["小时", "均值 (A)", "最大值 (A)", "最小值 (A)", "数据点数"], CLR_HEADER2)
    cur_row += 1
    h_rows = hourly_stats(data_running)
    if not h_rows:
        ws.cell(row=cur_row, column=1, value="本周无运行数据").alignment = _left()
        cur_row += 1
    else:
        for hi, hr in enumerate(h_rows):
            vals = [hr["小时"], hr["均值 (A)"], hr["最大值 (A)"], hr["最小值 (A)"], hr["数据点数"]]
            _write_data_row(ws, cur_row, vals, alt=(hi % 2 == 1))
            cur_row += 1

    cur_row += 1

    # ── 原始采样数据（运行数据，限 2000 行避免文件过大） ──
    MAX_RAW = 2000
    raw_data = data_running[:MAX_RAW]
    note_suffix = f"（仅显示前 {MAX_RAW} 条，共 {len(data_running)} 条）" if len(data_running) > MAX_RAW else f"（共 {len(data_running)} 条）"
    _section_title(ws, cur_row, f"📋  运行采样明细 {note_suffix}", 2)
    cur_row += 1
    _write_header_row(ws, cur_row, ["时间戳", "电流 (A)"], CLR_HEADER2)
    cur_row += 1
    for ri, d in enumerate(raw_data):
        c1 = ws.cell(row=cur_row, column=1, value=d["timestamp"][:19])
        c2 = ws.cell(row=cur_row, column=2, value=round(d["value"], 3))
        bg = CLR_ALT if ri % 2 == 1 else CLR_WHITE
        c1.fill = c2.fill = _fill(bg)
        c1.font = c2.font = _font()
        c1.alignment = _center()
        c2.alignment = _center()
        c1.border = c2.border = _border()
        cur_row += 1

    # ── 列宽 ──
    col_widths = [22, 12, 12, 12, 10]
    for ci, w in enumerate(col_widths, start=1):
        _set_col_width(ws, ci, w)


# ── 主函数 ────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="磺化釜电机电流 Excel 周报")
    parser.add_argument("--this-week", action="store_true", help="分析本周（周一~今天）")
    parser.add_argument("--start", type=str, default=None, help="开始日期 YYYY-MM-DD")
    parser.add_argument("--end",   type=str, default=None, help="结束日期 YYYY-MM-DD（含）")
    args = parser.parse_args()

    now = datetime.now()

    if args.start and args.end:
        start_dt = datetime.fromisoformat(args.start)
        end_dt   = datetime.fromisoformat(args.end) + timedelta(days=1)
        period_desc = f"{args.start} ~ {args.end}"
    elif args.this_week:
        # 本周周一到今天
        days_since_mon = now.weekday()  # 0=周一
        start_dt = (now - timedelta(days=days_since_mon)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        end_dt   = now
        period_desc = f"{start_dt.strftime('%Y-%m-%d')} ~ {now.strftime('%Y-%m-%d')}（本周）"
    else:
        # 上一完整周：上周周一 ~ 上周日
        days_since_mon = now.weekday()
        this_mon = (now - timedelta(days=days_since_mon)).replace(
            hour=0, minute=0, second=0, microsecond=0)
        last_mon = this_mon - timedelta(weeks=1)
        last_sun = this_mon  # 独占上限即本周一 00:00
        start_dt = last_mon
        end_dt   = last_sun
        period_desc = (
            f"{last_mon.strftime('%Y-%m-%d')} ~ "
            f"{(last_sun - timedelta(seconds=1)).strftime('%Y-%m-%d')}（上周）"
        )

    start_time    = start_dt.strftime("%Y-%m-%dT%H:%M:%S")
    end_time_excl = end_dt.strftime("%Y-%m-%dT%H:%M:%S")
    generated_at  = now.strftime("%Y-%m-%d %H:%M:%S")

    print("=" * 60)
    print(f"磺化釜电机电流周报 (Excel)")
    print(f"分析周期: {period_desc}")
    print(f"生成时间: {generated_at}")
    print("=" * 60)

    db = sqlite3.connect(DB_PATH)

    results = []
    for node_id in MOTOR_NODES:
        name = display_name(node_id)
        print(f"  读取 {name}...", end=" ", flush=True)
        data_all = read_data(db, node_id, start_time, end_time_excl)

        # 稳态数据提取（三阶段状态机，与预测报告一致）
        data_running = extract_steady_segments(data_all)
        stats = compute_stats(data_all, data_running)
        results.append({
            "node_id":            node_id,
            "data_all":           data_all,
            "data_running":       data_running,
            "stats":              stats,
        })
        print(f"全量 {stats['total']} 点，运行 {stats['count']} 点，"
              f"均值 {stats['avg']} A，运行 {stats['runtime_h']} h")

    db.close()

    # ── 创建 Workbook ──
    wb = Workbook()
    # 删除默认 Sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    build_summary_sheet(wb, results, period_desc, generated_at)
    for r in results:
        print(f"  生成 Sheet: {display_name(r['node_id'])}")
        build_motor_sheet(wb, r)

    # ── 保存 ──
    os.makedirs(REPORT_DIR, exist_ok=True)
    fname = f"motor_current_weekly_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    fpath = os.path.join(REPORT_DIR, fname)
    wb.save(fpath)

    print(f"\n✅ Excel 报告已生成：{fpath}")
    return fpath


if __name__ == "__main__":
    main()

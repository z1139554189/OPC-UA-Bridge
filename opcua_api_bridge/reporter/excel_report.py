"""
Excel 报表生成器（简洁版）
每分钟追加一行：时间戳 + 各节点实时值
布局：A 列时间，B~F 列对应 5 个节点的 PV 值
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path
from typing import Optional
import logging

logger = logging.getLogger(__name__)

# ── 辅助函数 ───────────────────────────────────────

def _short_node(node_id: str) -> str:
    """ns=1;s=FIT_05R301F01.PV -> FIT_05R301F01.PV"""
    if ";" in node_id:
        node_id = node_id.split(";", 1)[1]
    for prefix in ("s=", "i=", "g=", "b="):
        if node_id.startswith(prefix):
            node_id = node_id[len(prefix):]
            break
    return node_id


def _fmt_val(val) -> str:
    """数值保留两位小数，null/Bad 显示为空"""
    if val is None:
        return ""
    try:
        return str(round(float(val), 2))
    except (ValueError, TypeError):
        return ""


# ── 样式常量 ───────────────────────────────────────

COLOR_HEADER_BG   = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALT_ROW     = "EBF3FB"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_style(cell):
    cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


# ── 主入口 ────────────────────────────────────────

def generate_excel(
    nodes: list[str],
    values: list[dict],
    output_path: Optional[str] = None,
) -> str:
    """
    每分钟追加一行数据到 Excel。

    参数:
        nodes       - 节点 ID 列表（决定列头）
        values      - batch-read 返回的节点数据列表
        output_path - 输出文件路径

    返回:
        生成的文件绝对路径
    """
    if output_path is None:
        output_path = "opcua_report.xlsx"
    output_path = str(Path(output_path).resolve())

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    short_names = [_short_node(n) for n in nodes]

    if Path(output_path).exists():
        # ── 追加模式：在末尾加一行 ──
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        next_row = ws.max_row + 1

        # A 列：时间
        ws.cell(row=next_row, column=1, value=timestamp)
        ws.cell(row=next_row, column=1).border = THIN_BORDER
        ws.cell(row=next_row, column=1).alignment = Alignment(horizontal="center")

        # B~F 列：各节点值
        value_lookup = {v.get("node_id", ""): v for v in values}
        for col_i, node_id in enumerate(nodes, start=2):
            node_data = value_lookup.get(node_id, {})
            raw_val = node_data.get("value")
            quality = node_data.get("quality", "")
            # Bad 质量或 None 值留空
            if quality in ("Bad", "Error") or raw_val is None:
                cell_val = ""
            else:
                cell_val = _fmt_val(raw_val)

            cell = ws.cell(row=next_row, column=col_i, value=cell_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

            # 隔行着色
            if next_row % 2 == 0:
                ws.cell(row=next_row, column=1).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)
                cell.fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

        wb.save(output_path)
        logger.info(f"追加成功: {output_path} (第 {next_row - 1} 行)")
        return output_path

    # ── 新建模式 ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "实时数据"

    # 表头行
    headers = ["时间"] + short_names
    col_widths = [20] + [22] * len(short_names)
    for col_i, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_i, value=header)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # 第一行数据
    ws.cell(row=2, column=1, value=timestamp)
    ws.cell(row=2, column=1).border = THIN_BORDER
    ws.cell(row=2, column=1).alignment = Alignment(horizontal="center")

    value_lookup = {v.get("node_id", ""): v for v in values}
    for col_i, node_id in enumerate(nodes, start=2):
        node_data = value_lookup.get(node_id, {})
        raw_val = node_data.get("value")
        quality = node_data.get("quality", "")
        if quality in ("Bad", "Error") or raw_val is None:
            cell_val = ""
        else:
            cell_val = _fmt_val(raw_val)

        cell = ws.cell(row=2, column=col_i, value=cell_val)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

    # 冻结首行
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    wb.save(output_path)
    logger.info(f"报表已创建: {output_path}")
    print(f"[OK] 报表已创建: {output_path}")
    return output_path

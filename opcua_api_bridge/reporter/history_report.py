"""
历史数据报表生成器（一次性执行）
从 SQLite 历史库读取过去 1 小时 5 个 PV 节点的数据，生成 Excel 报表。
数据横向（节点列），时间纵向（行），每分钟一条。
"""

import sqlite3
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

# ── 配置 ──────────────────────────────────────────

DB_PATH = Path(__file__).parent.parent / "data" / "history.db"
OUTPUT_DIR = Path(__file__).parent / "output"

NODES = [
    "FIT_05R301F01.PV",
    "FIT_05R302F01.PV",
    "FIT_05R303F01.PV",
    "FIT_05R304F01.PV",
    "FIT_05R305F01.PV",
]

FULL_NODE_IDS = [f"ns=1;s={n}" for n in NODES]
HOURS_BACK = 1  # 回溯小时数


# ── 辅助函数 ──────────────────────────────────────

def _safe_table_name(node_id: str) -> str:
    """将 node_id 转为 SQLite 表名（与 HistoryDB._safe_table_name 一致）"""
    safe = node_id.replace("=", "_eq_").replace(";", "_sc_").replace(".", "_dot_")
    return f"h_{safe}"


def _fmt_val(val) -> str:
    """数值保留两位小数"""
    if val is None:
        return ""
    try:
        return str(round(float(val), 2))
    except (ValueError, TypeError):
        return ""


def _fmt_time(ts_str: str) -> str:
    """
    2026-04-02T13:24:25.289154 -> 2026-04-02 13:24:25
    去掉毫秒，精确到秒
    """
    if not ts_str:
        return ""
    try:
        dt = datetime.fromisoformat(ts_str)
        return dt.strftime("%Y-%m-%d %H:%M:%S")
    except ValueError:
        return ts_str


# ── 样式常量 ──────────────────────────────────────

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FONT = "FFFFFF"
COLOR_ALT_ROW = "EBF3FB"

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _header_style(cell):
    cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=11)
    cell.fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = THIN_BORDER


# ── 主逻辑 ────────────────────────────────────────

def main():
    if not DB_PATH.exists():
        print(f"[ERROR] SQLite not found: {DB_PATH}")
        sys.exit(1)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    end_time = datetime.now()
    start_time = end_time - timedelta(hours=HOURS_BACK)

    print(f"时间范围: {_fmt_time(start_time.isoformat())} ~ {_fmt_time(end_time.isoformat())}")
    print(f"节点数: {len(NODES)}")
    print(f"数据库: {DB_PATH}")
    print()

    conn = sqlite3.connect(str(DB_PATH))
    conn.row_factory = sqlite3.Row

    # ── 查询每个节点的历史数据 ──
    all_data = {}  # { node_short: [(timestamp_str, value), ...] }

    for node_id, node_short in zip(FULL_NODE_IDS, NODES):
        table_name = _safe_table_name(node_id)

        # 检查表是否存在
        cursor = conn.execute(
            "SELECT name FROM sqlite_master WHERE type='table' AND name=?",
            (table_name,),
        )
        if not cursor.fetchone():
            print(f"[WARN] {node_short}: 无历史数据表")
            all_data[node_short] = []
            continue

        sql = f"""
            SELECT timestamp, value, quality
            FROM [{table_name}]
            WHERE timestamp >= ? AND timestamp <= ?
            ORDER BY timestamp ASC
        """
        cursor = conn.execute(sql, (start_time.isoformat(), end_time.isoformat()))
        rows = cursor.fetchall()

        records = []
        for row in rows:
            ts = _fmt_time(row["timestamp"])
            val = row["value"]
            quality = row["quality"]
            if quality in ("Bad", "Error") or val is None:
                records.append((ts, None))
            else:
                records.append((ts, round(float(val), 2)))

        all_data[node_short] = records
        print(f"{node_short}: {len(records)} 条记录, {records[0][0] if records else 'N/A'} ~ {records[-1][0] if records else 'N/A'}")

    conn.close()
    print()

    # ── 采样：每分钟取一个值（取该分钟内最后一条） ──
    # 把所有节点的数据按分钟归并，确保每个时间行有完整的 5 个节点值
    minute_data = {}  # "2026-04-02 13:24" -> { node_short: value_or_None }

    for node_short, records in all_data.items():
        for ts_str, val in records:
            # 截取到分钟: "2026-04-02 13:24:25" -> "2026-04-02 13:24"
            minute_key = ts_str[:16]
            if minute_key not in minute_data:
                minute_data[minute_key] = {}
            # 同一分钟内取最后一个值
            minute_data[minute_key][node_short] = val

    # 按时间排序
    sorted_minutes = sorted(minute_data.keys())

    if not sorted_minutes:
        print("[ERROR] 过去 1 小时无任何数据")
        sys.exit(1)

    # ── 生成 Excel ──
    output_path = OUTPUT_DIR / f"opcua_history_{end_time.strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "历史数据"

    # 表头
    headers = ["时间"] + NODES
    col_widths = [20] + [22] * len(NODES)
    for col_i, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_i, value=header)
        _header_style(cell)
        ws.column_dimensions[get_column_letter(col_i)].width = width

    # 数据行
    for row_i, minute_key in enumerate(sorted_minutes, start=2):
        row_data = minute_data[minute_key]

        # A 列：时间（分钟 + ":00" 表示整分钟）
        display_time = minute_key + ":00"
        cell = ws.cell(row=row_i, column=1, value=display_time)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center")

        # B~F 列：各节点值
        for col_i, node_short in enumerate(NODES, start=2):
            val = row_data.get(node_short)
            cell_val = _fmt_val(val) if val is not None else ""
            cell = ws.cell(row=row_i, column=col_i, value=cell_val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center")

        # 隔行着色
        if row_i % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_i, column=col).fill = PatternFill("solid", fgColor=COLOR_ALT_ROW)

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(sorted_minutes) + 1}"

    # ── 添加折线图 ──
    data_rows = len(sorted_minutes)
    if data_rows >= 2:
        chart = LineChart()
        chart.title = f"过去 {HOURS_BACK} 小时 PV 趋势"
        chart.style = 10
        chart.y_axis.title = "PV 值"
        chart.x_axis.title = "时间"
        chart.width = 30
        chart.height = 15

        # X 轴：时间列
        x_data = Reference(ws, min_col=1, min_row=2, max_row=data_rows + 1)

        for col_i, node_short in enumerate(NODES, start=2):
            y_data = Reference(ws, min_col=col_i, min_row=1, max_row=data_rows + 1)
            chart.add_data(y_data, titles_from_data=True)

        chart.set_categories(x_data)

        # 设置线条颜色
        colors = ["FF6B6B", "4ECDC4", "45B7D1", "FFA07A", "98D8C8"]
        for i, series in enumerate(chart.series):
            series.graphicalProperties.line.width = 20000  # 线宽
            if i < len(colors):
                series.graphicalProperties.line.solidFill = colors[i]

        ws.add_chart(chart, f"A{data_rows + 3}")

    wb.save(str(output_path))

    total_rows = len(sorted_minutes)
    print(f"=" * 50)
    print(f"[OK] 历史报表已生成")
    print(f"  文件: {output_path}")
    print(f"  时间范围: {sorted_minutes[0]}:00 ~ {sorted_minutes[-1]}:00")
    print(f"  数据行数: {total_rows} 行（每分钟一条）")
    print(f"  节点数: {len(NODES)}")


if __name__ == "__main__":
    main()

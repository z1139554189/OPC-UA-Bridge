"""
实时采集并生成报表（持续采集版）
每 2 秒采集一次，持续 5 分钟，生成趋势报表

用法：
    python run_realtime_report.py
    python run_realtime_report.py --duration 10 --interval 1  # 10分钟，每秒采集
"""

import argparse
import sys
import logging
import time
from datetime import datetime, timezone, timedelta
from pathlib import Path
import threading

# 将项目根目录加入 path
ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(ROOT))

from reporter.collector import OpcUaCollector
from reporter.config import REPORT_CONFIG

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
logger = logging.getLogger(__name__)


def collect_realtime_data(
    collector: OpcUaCollector,
    nodes: list[str],
    duration_minutes: float,
    interval_seconds: float,
) -> dict:
    """
    持续采集实时数据
    
    参数:
        collector         - 采集器实例
        nodes             - 要采集的节点列表
        duration_minutes  - 采集持续时间（分钟）
        interval_seconds  - 采集间隔（秒）
    
    返回:
        {
            "start_time": "ISO时间",
            "end_time": "ISO时间",
            "interval_seconds": 2.0,
            "samples": [
                {
                    "timestamp": "ISO时间",
                    "values": [{"node_id": ..., "value": ..., "quality": ...}, ...]
                },
                ...
            ]
        }
    """
    result = {
        "start_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "interval_seconds": interval_seconds,
        "samples": [],
    }
    
    total_seconds = duration_minutes * 60
    total_samples = int(total_seconds / interval_seconds)
    
    logger.info(f"开始实时采集：{duration_minutes} 分钟，每 {interval_seconds} 秒采集一次")
    logger.info(f"预计采集 {total_samples} 个样本")
    
    collected = 0
    start_time = time.time()
    
    while collected < total_samples:
        loop_start = time.time()
        
        try:
            # 批量读取当前值
            batch_result = collector.batch_read(nodes)
            values = batch_result.get("results", [])
            
            # 时间戳：本地时间，精确到秒
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            sample = {
                "timestamp": now,
                "values": values,
            }
            result["samples"].append(sample)
            
            collected += 1
            if collected % 10 == 0 or collected == total_samples:
                logger.info(f"已采集 {collected}/{total_samples} 个样本")
                
        except Exception as e:
            logger.warning(f"第 {collected + 1} 次采集失败: {e}")
            result["samples"].append({
                "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "error": str(e),
            })
            collected += 1
        
        # 精确控制间隔
        elapsed = time.time() - loop_start
        sleep_time = interval_seconds - elapsed
        if sleep_time > 0:
            time.sleep(sleep_time)
    
    result["end_time"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    logger.info(f"采集完成！共 {len(result['samples'])} 个样本")
    
    return result


def generate_realtime_excel(data: dict, output_path: str = None) -> str:
    """
    生成实时趋势报表（Excel）
    
    参数:
        data          - 采集的数据
        output_path   - 输出路径（可选）
    
    返回:
        生成的文件路径
    """
    from openpyxl import Workbook
    from openpyxl.chart import LineChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    # 创建输出目录
    output_dir = Path(REPORT_CONFIG["output_dir"])
    output_dir.mkdir(parents=True, exist_ok=True)
    
    # 生成文件名
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = output_dir / f"realtime_report_{timestamp}.xlsx"
    else:
        output_path = Path(output_path)
    
    wb = Workbook()
    
    # ── Sheet 1: 摘要 ──
    ws_summary = wb.active
    ws_summary.title = "摘要"
    
    # 标题样式
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    
    ws_summary["A1"] = "OPC UA 实时数据报表"
    ws_summary["A1"].font = title_font
    ws_summary.merge_cells("A1:D1")
    
    ws_summary["A3"] = "采集信息"
    ws_summary["A3"].font = header_font
    
    info_rows = [
        ("开始时间", data.get("start_time", "")),
        ("结束时间", data.get("end_time", "")),
        ("采集间隔", f"{data.get('interval_seconds', 0)} 秒"),
        ("样本总数", str(len(data.get("samples", [])))),
    ]
    
    for i, (label, value) in enumerate(info_rows, start=4):
        ws_summary[f"A{i}"] = label
        ws_summary[f"B{i}"] = value
        ws_summary[f"A{i}"].font = header_font
    
    ws_summary.column_dimensions["A"].width = 20
    ws_summary.column_dimensions["B"].width = 40
    
    # ── Sheet 2: 实时数据 ──
    ws_data = wb.create_sheet("实时数据")
    
    # 提取节点列表
    samples = data.get("samples", [])
    if samples and "values" in samples[0]:
        first_sample = samples[0]["values"]
        node_ids = [v.get("node_id", "unknown") for v in first_sample]
    else:
        node_ids = []
    
    # 表头
    ws_data["A1"] = "时间戳"
    ws_data["A1"].font = header_font_white
    ws_data["A1"].fill = header_fill
    
    for col, node_id in enumerate(node_ids, start=2):
        cell = ws_data.cell(row=1, column=col, value=node_id)
        cell.font = header_font_white
        cell.fill = header_fill
    
    # 数据行
    for row, sample in enumerate(samples, start=2):
        if "error" in sample:
            ws_data.cell(row=row, column=1, value=sample.get("timestamp", ""))
            ws_data.cell(row=row, column=2, value=f"ERROR: {sample['error']}")
        else:
            ws_data.cell(row=row, column=1, value=sample.get("timestamp", ""))
            values = sample.get("values", [])
            for col, val_data in enumerate(values, start=2):
                value = val_data.get("value")
                quality = val_data.get("quality", "Unknown")
                if value is not None:
                    # 数值保留 2 位小数
                    if isinstance(value, float):
                        value = round(value, 2)
                    ws_data.cell(row=row, column=col, value=value)
                else:
                    ws_data.cell(row=row, column=col, value=f"[{quality}]")
    
    # 调整列宽
    ws_data.column_dimensions["A"].width = 30
    for col in range(2, len(node_ids) + 2):
        ws_data.column_dimensions[get_column_letter(col)].width = 25
    
    # ── Sheet 3: 趋势图 ──
    ws_chart = wb.create_sheet("趋势图")
    
    # 复制数据到趋势图 Sheet（用于图表引用）
    ws_chart["A1"] = "序号"
    ws_chart["A1"].font = header_font_white
    ws_chart["A1"].fill = header_fill
    
    for col, node_id in enumerate(node_ids, start=2):
        cell = ws_chart.cell(row=1, column=col, value=node_id)
        cell.font = header_font_white
        cell.fill = header_fill
    
    # 填充数据
    for row, sample in enumerate(samples, start=2):
        seq = row - 1
        ws_chart.cell(row=row, column=1, value=seq)
        if "values" in sample:
            for col, val_data in enumerate(sample["values"], start=2):
                value = val_data.get("value")
                if value is not None and isinstance(value, (int, float)):
                    if isinstance(value, float):
                        value = round(value, 2)
                    ws_chart.cell(row=row, column=col, value=value)
    
    # 创建折线图
    if node_ids:
        chart = LineChart()
        chart.title = "实时趋势"
        chart.style = 13
        chart.y_axis.title = "值"
        chart.x_axis.title = "样本序号"
        chart.width = 20
        chart.height = 12
        
        # 数据范围
        data_end_row = len(samples) + 1
        data_end_col = len(node_ids) + 1
        
        data_ref = Reference(
            ws_chart,
            min_col=2,
            min_row=1,
            max_col=data_end_col,
            max_row=data_end_row,
        )
        cats_ref = Reference(
            ws_chart,
            min_col=1,
            min_row=2,
            max_row=data_end_row,
        )
        
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        
        # 添加到 Sheet
        ws_chart.add_chart(chart, "A" + str(data_end_row + 3))
    
    # 保存文件
    wb.save(output_path)
    return str(output_path)


def main():
    parser = argparse.ArgumentParser(description="OPC UA 实时数据采集 & Excel 报表生成")
    parser.add_argument(
        "--url",
        default=REPORT_CONFIG["api_url"],
        help="桥接器 API 地址，默认 http://localhost:8000"
    )
    parser.add_argument(
        "--duration",
        type=float,
        default=5.0,
        help="采集持续时间（分钟），默认 5 分钟"
    )
    parser.add_argument(
        "--interval",
        type=float,
        default=2.0,
        help="采集间隔（秒），默认 2 秒"
    )
    parser.add_argument(
        "--out",
        default=None,
        help="输出文件路径，默认自动生成带时间戳文件名"
    )
    args = parser.parse_args()
    
    collector = OpcUaCollector(base_url=args.url)
    
    # ── 先检查服务是否可达 ──
    print(f"[INFO] 连接桥接器: {args.url}")
    try:
        health = collector.check_health()
        opc_ok = health.get("opcua_connected", False)
        print(f"   健康状态: {'[OK] OPC UA 已连接' if opc_ok else '[WARN] OPC UA 未连接'}")
    except Exception as e:
        print(f"[ERROR] 无法连接桥接器: {e}")
        print("   请先确认桥接器已启动")
        sys.exit(1)
    
    # ── 持续采集 ──
    nodes = REPORT_CONFIG["fixed_nodes"]
    print(f"[INFO] 采集节点: {nodes}")
    
    data = collect_realtime_data(
        collector=collector,
        nodes=nodes,
        duration_minutes=args.duration,
        interval_seconds=args.interval,
    )
    
    # ── 生成报表 ──
    print("[INFO] 生成 Excel 报表...")
    out_path = generate_realtime_excel(data, output_path=args.out)
    print(f"\n[DONE] 完成！文件已保存至:\n   {out_path}")


if __name__ == "__main__":
    main()
